from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from sqlalchemy import func, select

from app.api.routes import imports as import_routes
from app.core.config import get_settings
from app.core.security import hash_password
from app.models import AuditLog, ImportBatch, ImportedWeeklyPlanRaw, ImportRowIssue, InventorySnapshot, PipeWipSnapshot, Product, RegularProductionProduct, Role, User, UserRole, WeeklyPlanStagingRow


def workbook_bytes(headers, rows, *, sheet_name="数据", extra_sheet=False, multiline=False, trailing_style_row=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    if multiline:
        sheet.merge_cells("A1:A2")
        sheet["A1"] = headers[0]
        sheet.merge_cells("B1:C1")
        sheet["B1"] = "产品信息"
        sheet["B2"] = headers[1]
        sheet["C2"] = headers[2]
        header_end = 2
    else:
        sheet.append(headers)
        header_end = 1
    for row in rows:
        sheet.append(row)
    if trailing_style_row:
        sheet.cell(trailing_style_row, 1).fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    if extra_sheet:
        workbook.create_sheet("说明")["A1"] = "虚拟测试说明"
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue(), header_end


def upload(client, token, content, import_type="INVENTORY", filename="synthetic.xlsx", **data):
    form = {"import_type": import_type, **data}
    return client.post(
        "/api/v1/imports/upload",
        headers={"Authorization": f"Bearer {token}"},
        data=form,
        files={"file": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


def analyze(client, token, batch_id, sheet_name="数据", header_end=None):
    payload = {"sheet_name": sheet_name}
    if header_end:
        payload.update({"header_row_start": 1, "header_row_end": header_end})
    return client.post(
        f"/api/v1/imports/{batch_id}/analyze",
        headers={"Authorization": f"Bearer {token}"},
        json=payload,
    )


def validate(client, token, batch_id):
    return client.post(
        f"/api/v1/imports/{batch_id}/validate",
        headers={"Authorization": f"Bearer {token}"},
    )


def test_standard_upload_analyze_validate_confirm_and_inventory_recalculation(client, db, admin_token):
    content, _ = workbook_bytes(
        ["产品编码", "产品名称", "规格型号", "现存数量", "预计入库", "预计出库", "可用数量"],
        [["00001234", "虚拟水管", "DN20", 10, "3", 2, 99]],
        extra_sheet=True,
    )
    response = upload(client, admin_token, content, source_date="2026-07-18")
    assert response.status_code == 201
    batch_id = response.json()["id"]
    assert response.json()["sheet_names"] == ["数据", "说明"]

    sheets = client.get(f"/api/v1/imports/{batch_id}/sheets", headers={"Authorization": f"Bearer {admin_token}"})
    assert sheets.status_code == 200
    assert sheets.json()["sheet_count"] == 2
    analyzed = analyze(client, admin_token, batch_id)
    assert analyzed.status_code == 200
    assert analyzed.json()["field_mapping"]["product_code"] == 1

    checked = validate(client, admin_token, batch_id)
    assert checked.status_code == 200
    assert checked.json()["status"] == "READY"
    assert checked.json()["warning_rows"] == 1

    preview = client.get(f"/api/v1/imports/{batch_id}/preview", headers={"Authorization": f"Bearer {admin_token}"})
    assert preview.json()["items"][0]["data"]["product_code"] == "00001234"
    assert preview.json()["items"][0]["data"]["calculated_available_qty"] == "11"

    confirmed = client.post(f"/api/v1/imports/{batch_id}/confirm", headers={"Authorization": f"Bearer {admin_token}"})
    assert confirmed.status_code == 200
    assert confirmed.json()["status"] == "COMPLETED"
    assert db.scalar(select(Product).where(Product.product_code == "00001234")) is not None
    snapshot = db.scalar(select(InventorySnapshot).where(InventorySnapshot.import_batch_id == batch_id))
    assert str(snapshot.calculated_available_qty) == "11.0000"
    assert str(snapshot.source_available_qty) == "99.0000"


def test_non_xlsx_and_oversized_files_are_rejected(client, admin_token, monkeypatch):
    rejected = upload(client, admin_token, b"plain text", filename="unsafe.csv")
    assert rejected.status_code == 415
    assert rejected.json()["code"] == "XLSX_REQUIRED"
    monkeypatch.setattr(get_settings(), "import_max_file_size_mb", 0)
    too_large = upload(client, admin_token, b"x", filename="large.xlsx")
    assert too_large.status_code == 413
    assert too_large.json()["code"] == "IMPORT_FILE_TOO_LARGE"


def test_duplicate_file_is_blocked_and_admin_may_force_with_reason(client, admin_token):
    content, _ = workbook_bytes(["产品编码"], [["TEST-PIPE-001"]])
    first = upload(client, admin_token, content, import_type="REGULAR_PRODUCT", source_date="2026-07-18")
    assert first.status_code == 201
    duplicate = upload(client, admin_token, content, import_type="REGULAR_PRODUCT", source_date="2026-07-18")
    assert duplicate.status_code == 409
    assert duplicate.json()["code"] == "DUPLICATE_IMPORT_FILE"
    forced = upload(client, admin_token, content, import_type="REGULAR_PRODUCT", source_date="2026-07-18", force="true", force_reason="重新核对映射")
    assert forced.status_code == 201


def test_multiline_header_is_detected(client, admin_token):
    content, _ = workbook_bytes(["产品编码", "产品名称", "规格型号"], [["TEST-PIPE-001", "虚拟水管", "DN20"]], multiline=True)
    batch_id = upload(client, admin_token, content, import_type="REGULAR_PRODUCT").json()["id"]
    result = analyze(client, admin_token, batch_id)
    assert result.status_code == 200
    assert result.json()["analysis"]["header_row_start"] == 1
    assert result.json()["analysis"]["header_row_end"] == 2
    assert result.json()["field_mapping"] == {"product_code": 1, "product_name": 2, "specification": 3}


def test_scientific_notation_product_code_is_an_error(client, admin_token):
    content, _ = workbook_bytes(["产品编码"], [["1.2345E+11"]])
    batch_id = upload(client, admin_token, content, import_type="REGULAR_PRODUCT").json()["id"]
    analyze(client, admin_token, batch_id)
    result = validate(client, admin_token, batch_id)
    assert result.json()["status"] == "VALIDATION_FAILED"
    issues = client.get(f"/api/v1/imports/{batch_id}/issues", headers={"Authorization": f"Bearer {admin_token}"}).json()["items"]
    assert "PRODUCT_CODE_SCIENTIFIC_NOTATION" in {item["issue_code"] for item in issues}


def test_negative_wip_is_warning_and_can_be_confirmed(client, db, admin_token):
    content, _ = workbook_bytes(["产品编码", "产品名称", "未入库数量"], [["TEST-PIPE-001", "虚拟水管", -5]])
    batch_id = upload(client, admin_token, content, import_type="PIPE_WIP", source_date="2026-07-18").json()["id"]
    analyze(client, admin_token, batch_id)
    result = validate(client, admin_token, batch_id)
    assert result.json()["status"] == "READY"
    assert result.json()["warning_rows"] == 1
    confirmed = client.post(f"/api/v1/imports/{batch_id}/confirm", headers={"Authorization": f"Bearer {admin_token}"})
    assert confirmed.status_code == 200
    snapshot = db.scalar(select(PipeWipSnapshot).where(PipeWipSnapshot.import_batch_id == batch_id))
    assert str(snapshot.quantity) == "-5.0000"


def test_product_identity_conflict_and_duplicate_are_reported(client, admin_token):
    content, _ = workbook_bytes(
        ["产品编码", "产品名称", "规格", "未入库数量"],
        [["TEST-PIPE-001", "名称一", "DN20", 1], ["TEST-PIPE-001", "名称二", "DN25", 2]],
    )
    batch_id = upload(client, admin_token, content, import_type="PIPE_WIP", source_date="2026-07-18").json()["id"]
    analyze(client, admin_token, batch_id)
    validate(client, admin_token, batch_id)
    issues = client.get(f"/api/v1/imports/{batch_id}/issues", headers={"Authorization": f"Bearer {admin_token}"}).json()["items"]
    codes = {item["issue_code"] for item in issues}
    assert "PRODUCT_IDENTITY_CONFLICT" in codes
    assert "DUPLICATE_ROW" in codes


def test_error_rows_cannot_be_confirmed_or_enter_business_tables(client, db, admin_token):
    content, _ = workbook_bytes(["产品编码", "现存数量", "预计入库", "预计出库"], [["TEST-FITTING-001", "bad", 0, 0]])
    batch_id = upload(client, admin_token, content).json()["id"]
    analyze(client, admin_token, batch_id)
    validate(client, admin_token, batch_id)
    response = client.post(f"/api/v1/imports/{batch_id}/confirm", headers={"Authorization": f"Bearer {admin_token}"})
    assert response.status_code == 409
    assert response.json()["code"] == "IMPORT_NOT_READY"
    assert db.scalar(select(func.count(InventorySnapshot.id))) == 0


def test_transaction_failure_rolls_back_and_marks_batch_failed(client, db, admin_token, monkeypatch):
    content, _ = workbook_bytes(["产品编码"], [["TEST-CLAMP-108"]])
    batch_id = upload(client, admin_token, content, import_type="REGULAR_PRODUCT").json()["id"]
    analyze(client, admin_token, batch_id)
    validate(client, admin_token, batch_id)
    monkeypatch.setattr(import_routes, "import_validated_batch", lambda *args: (_ for _ in ()).throw(RuntimeError("synthetic failure")))
    response = client.post(f"/api/v1/imports/{batch_id}/confirm", headers={"Authorization": f"Bearer {admin_token}"})
    assert response.status_code == 500
    db.expire_all()
    assert db.get(ImportBatch, batch_id).status == "FAILED"
    assert db.scalar(select(func.count(Product.id)).where(Product.product_code == "TEST-CLAMP-108")) == 0


def completed_regular_batch(client, admin_token):
    content, _ = workbook_bytes(["产品编码"], [["TEST-CLAMP-108"]])
    batch_id = upload(client, admin_token, content, import_type="REGULAR_PRODUCT").json()["id"]
    analyze(client, admin_token, batch_id)
    validate(client, admin_token, batch_id)
    assert client.post(f"/api/v1/imports/{batch_id}/confirm", headers={"Authorization": f"Bearer {admin_token}"}).status_code == 200
    return batch_id


def test_completed_batch_can_be_rolled_back_with_audit(client, db, admin_token):
    batch_id = completed_regular_batch(client, admin_token)
    response = client.post(f"/api/v1/imports/{batch_id}/rollback", headers={"Authorization": f"Bearer {admin_token}"}, json={"reason": "撤销虚拟测试导入"})
    assert response.status_code == 200
    assert response.json()["status"] == "ROLLED_BACK"
    audit = db.scalar(select(AuditLog).where(AuditLog.action == "import.rollback", AuditLog.entity_id == str(batch_id)))
    assert audit.before_data["status"] == "COMPLETED"
    assert audit.after_data["status"] == "ROLLED_BACK"
    assert audit.reason == "撤销虚拟测试导入"


def test_referenced_batch_cannot_be_rolled_back(client, admin_token, monkeypatch):
    batch_id = completed_regular_batch(client, admin_token)
    monkeypatch.setattr(import_routes, "batch_has_downstream_references", lambda *args: True)
    response = client.post(f"/api/v1/imports/{batch_id}/rollback", headers={"Authorization": f"Bearer {admin_token}"}, json={"reason": "尝试撤销"})
    assert response.status_code == 409
    assert response.json()["code"] == "IMPORT_BATCH_REFERENCED"


def test_viewer_can_list_but_cannot_upload(client, db, admin_token):
    viewer_role = db.scalar(select(Role).where(Role.code == "VIEWER"))
    viewer = User(username="viewer-import", display_name="只读测试用户", password_hash=hash_password("ViewerTest123!"))
    viewer.role_links.append(UserRole(role=viewer_role))
    db.add(viewer)
    db.commit()
    login = client.post("/api/v1/auth/login", json={"username": "viewer-import", "password": "ViewerTest123!"})
    token = login.json()["access_token"]
    assert client.get("/api/v1/imports", headers={"Authorization": f"Bearer {token}"}).status_code == 200
    content, _ = workbook_bytes(["产品编码"], [["TEST-PIPE-001"]])
    assert upload(client, token, content, import_type="REGULAR_PRODUCT").status_code == 403


def test_invalid_workbook_response_does_not_leak_server_path(client, admin_token):
    response = upload(client, admin_token, b"not a zip archive")
    assert response.status_code == 422
    body = response.json()
    assert body["code"] == "INVALID_WORKBOOK"
    assert "F:\\" not in str(body) and "/data/" not in str(body)


def test_effective_range_stops_before_million_formatted_rows(client, admin_token):
    content, _ = workbook_bytes(["产品编码"], [["TEST-PIPE-001"]], trailing_style_row=1_048_576)
    batch_id = upload(client, admin_token, content, import_type="REGULAR_PRODUCT").json()["id"]
    result = analyze(client, admin_token, batch_id)
    analysis = result.json()["analysis"]
    assert analysis["declared_rows"] == 1_048_576
    assert analysis["last_row"] == 2
    assert analysis["scanned_rows"] < 500


def test_upload_analyze_validate_and_confirm_are_audited(client, db, admin_token):
    batch_id = completed_regular_batch(client, admin_token)
    actions = set(db.scalars(select(AuditLog.action).where(AuditLog.entity_id == str(batch_id))).all())
    assert {"import.upload", "import.analyze", "import.validate", "import.confirm"}.issubset(actions)


def hidden_rows_workbook():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["产品编码"])
    sheet.append(["TEST-PIPE-001"])
    sheet.append(["TEST-FITTING-001"])
    sheet.row_dimensions[3].hidden = True
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_hidden_rows_are_imported_by_default_with_metadata(client, db, admin_token):
    batch_id = upload(client, admin_token, hidden_rows_workbook(), import_type="REGULAR_PRODUCT").json()["id"]
    analyzed = analyze(client, admin_token, batch_id)
    assert analyzed.json()["analysis"]["hidden_row_count"] == 1
    checked = validate(client, admin_token, batch_id).json()
    assert checked["total_rows"] == 2
    assert checked["hidden_data_rows"] == 1
    assert checked["excluded_hidden_data_rows"] == 0
    client.post(f"/api/v1/imports/{batch_id}/confirm", headers={"Authorization": f"Bearer {admin_token}"})
    records = db.scalars(select(RegularProductionProduct).order_by(RegularProductionProduct.source_row_number)).all()
    assert len(records) == 2
    assert records[0].raw_data["source_row_hidden"] is False
    assert records[1].raw_data["source_row_hidden"] is True
    upload_audit = db.scalar(select(AuditLog).where(AuditLog.action == "import.upload", AuditLog.entity_id == str(batch_id)))
    assert upload_audit.after_data["include_hidden_rows"] is True


def test_visible_rows_only_excludes_hidden_rows_and_audits_statistics(client, db, admin_token):
    batch_id = upload(client, admin_token, hidden_rows_workbook(), import_type="REGULAR_PRODUCT", include_hidden_rows="false").json()["id"]
    analyze(client, admin_token, batch_id)
    checked = validate(client, admin_token, batch_id).json()
    assert checked["total_rows"] == 1
    assert checked["hidden_data_rows"] == 1
    assert checked["excluded_hidden_data_rows"] == 1
    client.post(f"/api/v1/imports/{batch_id}/confirm", headers={"Authorization": f"Bearer {admin_token}"})
    assert db.scalar(select(func.count(RegularProductionProduct.id))) == 1
    validation_audit = db.scalar(select(AuditLog).where(AuditLog.action == "import.validate", AuditLog.entity_id == str(batch_id)))
    assert validation_audit.after_data["hidden_data_rows"] == 1
    assert validation_audit.after_data["excluded_hidden_data_rows"] == 1


def realistic_weekly_plan_workbook():
    workbook = Workbook()
    workbook.remove(workbook.active)
    layouts = {
        "制管": {"header_end": 5, "equipment": 1, "name": 3, "spec": 4, "batch": 5, "process": 6, "kind": 12, "total": 13, "daily": 14, "start": "2026-07-13"},
        "包装": {"header_end": 4, "equipment": 1, "name": 2, "spec": 3, "batch": 4, "process": 5, "kind": 10, "total": 11, "daily": 12, "start": "2026-07-13"},
        "成型": {"header_end": 5, "equipment": 1, "name": 3, "spec": 4, "batch": 5, "process": 6, "kind": 11, "total": 12, "daily": 13, "start": "2026-07-13"},
        "下料": {"header_end": 4, "equipment": None, "name": 1, "spec": 2, "batch": 3, "process": None, "kind": 8, "total": 9, "daily": 10, "start": "2026-07-10"},
    }
    from datetime import date, timedelta

    for sheet_name, layout in layouts.items():
        sheet = workbook.create_sheet(sheet_name)
        start = date.fromisoformat(layout["start"])
        end = start + timedelta(days=6)
        sheet.cell(1, 1, f"{sheet_name}周计划 {start:%Y-%m-%d} - {end:%Y-%m-%d}")
        for index in range(7):
            sheet.cell(layout["header_end"], layout["daily"] + index, start + timedelta(days=index))
        plan_row = layout["header_end"] + 1
        actual_row = plan_row + 1
        values = {
            layout["name"]: "虚拟水管",
            layout["spec"]: "DN20",
            layout["batch"]: f"LOT-DEMO-{sheet_name}",
            layout["kind"]: "计划",
        }
        if layout["equipment"]:
            values[layout["equipment"]] = f"设备-{sheet_name}"
        if layout["process"]:
            values[layout["process"]] = sheet_name
        for column, value in values.items():
            sheet.cell(plan_row, column, value)
            if column != layout["kind"]:
                sheet.merge_cells(start_row=plan_row, start_column=column, end_row=actual_row, end_column=column)
        sheet.cell(actual_row, layout["kind"], "实际")
        for index in range(7):
            sheet.cell(plan_row, layout["daily"] + index, index + 1)
            if sheet_name == "制管":
                sheet.cell(actual_row, layout["daily"] + index, f"='[日报-DEMO.xlsx]日报'!A{index + 1}")
            else:
                sheet.cell(actual_row, layout["daily"] + index, index)
        sheet.cell(plan_row, layout["total"], f"=SUM({sheet.cell(plan_row, layout['daily']).coordinate}:{sheet.cell(plan_row, layout['daily'] + 6).coordinate})")
        sheet.cell(actual_row, layout["total"], f"=SUM({sheet.cell(actual_row, layout['daily']).coordinate}:{sheet.cell(actual_row, layout['daily'] + 6).coordinate})")
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_realistic_four_sheet_weekly_plan_stages_daily_values_and_requires_manual_match(client, db, admin_token):
    content = realistic_weekly_plan_workbook()
    batch_ids = []
    for index, sheet_name in enumerate(("制管", "包装", "成型", "下料")):
        response = upload(
            client, admin_token, content, import_type="WEEKLY_PLAN",
            force="true" if index else "false", force_reason="测试四类周计划工作表" if index else None,
        )
        assert response.status_code == 201, response.text
        batch_id = response.json()["id"]
        batch_ids.append(batch_id)
        analyzed = analyze(client, admin_token, batch_id, sheet_name=sheet_name)
        assert analyzed.status_code == 200, analyzed.text
        checked = validate(client, admin_token, batch_id)
        assert checked.json()["status"] == "READY", checked.text
        confirmed = client.post(f"/api/v1/imports/{batch_id}/confirm", headers={"Authorization": f"Bearer {admin_token}"})
        assert confirmed.status_code == 200, confirmed.text
        assert confirmed.json()["imported_rows"] == 1

    assert db.scalar(select(func.count(WeeklyPlanStagingRow.id))) == 4
    assert db.scalar(select(func.count(ImportedWeeklyPlanRaw.id))) == 0
    assert db.scalar(select(func.count(Product.id))) == 0
    staging = db.scalar(select(WeeklyPlanStagingRow).where(WeeklyPlanStagingRow.import_batch_id == batch_ids[0]))
    assert staging.match_status == "UNMATCHED"
    assert len(staging.daily_plan) == 7
    assert staging.daily_plan["2026-07-13"] == "1"
    assert staging.daily_actual["2026-07-13"] is None
    assert staging.formula_metadata["external_actual_formulas"]

    product = Product(product_code="TEST-PIPE-001", product_name="虚拟水管", specification="DN20", data_source="TEST")
    db.add(product)
    db.commit()
    candidates = client.get(f"/api/v1/imports/{batch_ids[0]}/product-candidates", headers={"Authorization": f"Bearer {admin_token}"}, params={"keyword": "虚拟水管"})
    assert candidates.json()["items"][0]["product_code"] == "TEST-PIPE-001"
    matched = client.post(
        f"/api/v1/imports/{batch_ids[0]}/weekly-plan-staging/{staging.id}/match",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"action": "MATCH", "product_id": product.id, "reason": "根据虚拟编码人工确认"},
    )
    assert matched.status_code == 200
    assert matched.json()["match_status"] == "MATCHED"
    raw = db.scalar(select(ImportedWeeklyPlanRaw).where(ImportedWeeklyPlanRaw.import_batch_id == batch_ids[0]))
    assert raw.product_id == product.id
    assert raw.daily_plan["2026-07-19"] == "7"
    ignored = client.post(
        f"/api/v1/imports/{batch_ids[0]}/weekly-plan-staging/{staging.id}/match",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"action": "IGNORE", "reason": "人工复核后忽略虚拟行"},
    )
    assert ignored.status_code == 200
    assert ignored.json()["match_status"] == "IGNORED"
    assert db.scalar(select(ImportedWeeklyPlanRaw).where(ImportedWeeklyPlanRaw.import_batch_id == batch_ids[0])) is None


def test_duplicate_mapping_is_rejected(client, admin_token):
    content, _ = workbook_bytes(["产品编码", "产品名称"], [["TEST-PIPE-001", "虚拟水管"]])
    batch_id = upload(client, admin_token, content, import_type="REGULAR_PRODUCT").json()["id"]
    analyze(client, admin_token, batch_id)
    response = client.put(
        f"/api/v1/imports/{batch_id}/mapping",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"field_mapping": {"product_code": 1, "product_name": 1}, "conversion_rules": {}},
    )
    assert response.status_code == 422
    assert response.json()["code"] == "DUPLICATE_FIELD_MAPPING"


def test_issue_export_is_authenticated_csv(client, admin_token):
    content, _ = workbook_bytes(["产品编码"], [["1.2E+9"]])
    batch_id = upload(client, admin_token, content, import_type="REGULAR_PRODUCT").json()["id"]
    analyze(client, admin_token, batch_id)
    validate(client, admin_token, batch_id)
    assert client.get(f"/api/v1/imports/{batch_id}/issues/export").status_code == 401
    response = client.get(f"/api/v1/imports/{batch_id}/issues/export", headers={"Authorization": f"Bearer {admin_token}"})
    assert response.status_code == 200
    assert response.content.startswith(b"\xef\xbb\xbf")
    assert "PRODUCT_CODE_SCIENTIFIC_NOTATION" in response.content.decode("utf-8-sig")


def test_uploaded_filename_is_sanitized_and_server_path_is_never_returned(client, admin_token):
    content, _ = workbook_bytes(["产品编码"], [["TEST-PIPE-001"]])
    response = upload(client, admin_token, content, import_type="REGULAR_PRODUCT", filename="../../secret.xlsx")
    assert response.status_code == 201
    assert response.json()["original_filename"] == "secret.xlsx"
    assert "stored_filename" not in response.json()


def confirm_inventory(client, token, headers, row, *, source_date=None):
    content, _ = workbook_bytes(headers, [row])
    response = upload(client, token, content, source_date=source_date) if source_date else upload(client, token, content)
    batch_id = response.json()["id"]
    analyze(client, token, batch_id)
    checked = validate(client, token, batch_id)
    if checked.json()["status"] == "READY":
        client.post(f"/api/v1/imports/{batch_id}/confirm", headers={"Authorization": f"Bearer {token}"})
    return batch_id, checked


def test_snapshot_date_prefers_excel_value_over_batch_source_date(client, db, admin_token):
    batch_id, checked = confirm_inventory(
        client, admin_token,
        ["产品编码", "快照日期", "现存数量", "预计入库", "预计出库"],
        ["TEST-PIPE-001", "2026-07-17", 10, 0, 0],
        source_date="2026-07-18",
    )
    assert checked.json()["status"] == "READY"
    snapshot = db.scalar(select(InventorySnapshot).where(InventorySnapshot.import_batch_id == batch_id))
    assert snapshot.snapshot_date.isoformat() == "2026-07-17"


def test_snapshot_date_falls_back_to_source_date(client, db, admin_token):
    batch_id, checked = confirm_inventory(
        client, admin_token,
        ["产品编码", "现存数量", "预计入库", "预计出库"],
        ["TEST-PIPE-001", 10, 0, 0],
        source_date="2026-07-18",
    )
    assert checked.json()["status"] == "READY"
    snapshot = db.scalar(select(InventorySnapshot).where(InventorySnapshot.import_batch_id == batch_id))
    assert snapshot.snapshot_date.isoformat() == "2026-07-18"


def test_snapshot_date_is_required_when_excel_and_source_date_are_missing(client, db, admin_token):
    batch_id, checked = confirm_inventory(
        client, admin_token,
        ["产品编码", "现存数量", "预计入库", "预计出库"],
        ["TEST-PIPE-001", 10, 0, 0],
    )
    assert checked.json()["status"] == "VALIDATION_FAILED"
    issues = db.scalars(select(ImportRowIssue).where(ImportRowIssue.import_batch_id == batch_id)).all()
    assert "SNAPSHOT_DATE_REQUIRED" in {item.issue_code for item in issues}
    response = client.post(f"/api/v1/imports/{batch_id}/confirm", headers={"Authorization": f"Bearer {admin_token}"})
    assert response.status_code == 409


def test_snapshot_dates_remain_distinct_between_batches(client, db, admin_token):
    first_id, _ = confirm_inventory(client, admin_token, ["产品编码", "现存数量", "预计入库", "预计出库"], ["TEST-PIPE-001", 10, 0, 0], source_date="2026-07-17")
    second_id, _ = confirm_inventory(client, admin_token, ["产品编码", "现存数量", "预计入库", "预计出库"], ["TEST-PIPE-001", 11, 0, 0], source_date="2026-07-18")
    snapshots = db.scalars(select(InventorySnapshot).where(InventorySnapshot.import_batch_id.in_([first_id, second_id])).order_by(InventorySnapshot.import_batch_id)).all()
    assert [item.snapshot_date.isoformat() for item in snapshots] == ["2026-07-17", "2026-07-18"]


def import_regular_product(client, token, row, *, policy="FILL_EMPTY", admin_update_reason=None):
    content, _ = workbook_bytes(["产品编码", "产品名称", "规格", "分类", "单位"], [row])
    batch_id = upload(client, token, content, import_type="REGULAR_PRODUCT", master_data_policy="KEEP_EXISTING" if policy == "KEEP_EXISTING" else "FILL_EMPTY").json()["id"]
    analyze(client, token, batch_id)
    if policy == "ADMIN_UPDATE":
        response = client.put(
            f"/api/v1/imports/{batch_id}/options",
            headers={"Authorization": f"Bearer {token}"},
            json={"include_hidden_rows": True, "source_date": None, "master_data_policy": "ADMIN_UPDATE", "master_data_reason": admin_update_reason},
        )
        assert response.status_code == 200
    checked = validate(client, token, batch_id)
    assert checked.json()["status"] == "READY"
    confirmed = client.post(f"/api/v1/imports/{batch_id}/confirm", headers={"Authorization": f"Bearer {token}"})
    assert confirmed.status_code == 200
    return batch_id


def test_cross_batch_product_conflicts_fill_only_empty_fields_and_are_audited(client, db, admin_token):
    first_id = import_regular_product(client, admin_token, ["TEST-PIPE-001", "原名称", None, None, None])
    second_id = import_regular_product(client, admin_token, ["TEST-PIPE-001", "新名称", "DN20", "水管", "支"])
    db.expire_all()
    product = db.scalar(select(Product).where(Product.product_code == "TEST-PIPE-001"))
    assert product.product_name == "原名称"
    assert product.specification == "DN20"
    assert product.category == "水管"
    assert product.unit == "支"
    assert product.last_import_batch_id == second_id
    issues = db.scalars(select(ImportRowIssue).where(ImportRowIssue.import_batch_id == second_id)).all()
    name_issue = next(item for item in issues if item.issue_code == "PRODUCT_MASTER_NAME_CONFLICT")
    assert "新名称" in name_issue.raw_value and "原名称" in name_issue.raw_value
    audit = db.scalar(select(AuditLog).where(AuditLog.action == "product.master_data.import", AuditLog.entity_id == str(product.id)).order_by(AuditLog.id.desc()))
    assert set(audit.after_data["changed_fields"]) == {"specification", "category", "unit"}
    assert audit.after_data["import_batch_id"] == second_id
    assert first_id != second_id


def test_keep_existing_and_admin_update_product_master_policies(client, db, admin_token):
    import_regular_product(client, admin_token, ["TEST-FITTING-001", "原名称", "DN20", "管件", "只"])
    keep_id = import_regular_product(client, admin_token, ["TEST-FITTING-001", "冲突名称", "DN25", "新分类", "箱"], policy="KEEP_EXISTING")
    db.expire_all()
    product = db.scalar(select(Product).where(Product.product_code == "TEST-FITTING-001"))
    assert (product.product_name, product.specification, product.category, product.unit) == ("原名称", "DN20", "管件", "只")
    conflict_codes = set(db.scalars(select(ImportRowIssue.issue_code).where(ImportRowIssue.import_batch_id == keep_id)).all())
    assert {"PRODUCT_MASTER_NAME_CONFLICT", "PRODUCT_MASTER_SPEC_CONFLICT", "PRODUCT_MASTER_CATEGORY_CONFLICT", "PRODUCT_MASTER_UNIT_CONFLICT"}.issubset(conflict_codes)

    update_id = import_regular_product(client, admin_token, ["TEST-FITTING-001", "管理员确认名称", "DN25", "新分类", "箱"], policy="ADMIN_UPDATE", admin_update_reason="审查后确认更新虚拟主数据")
    db.expire_all()
    product = db.scalar(select(Product).where(Product.product_code == "TEST-FITTING-001"))
    assert (product.product_name, product.specification, product.category, product.unit) == ("管理员确认名称", "DN25", "新分类", "箱")
    assert product.last_import_batch_id == update_id
    audit = db.scalar(select(AuditLog).where(AuditLog.action == "product.master_data.import", AuditLog.entity_id == str(product.id)).order_by(AuditLog.id.desc()))
    assert audit.reason == "审查后确认更新虚拟主数据"


def test_issue_csv_escapes_formula_prefixes_but_preserves_negative_numbers_and_utf8(client, db, admin_token):
    content, _ = workbook_bytes(["产品编码"], [["TEST-PIPE-001"]])
    batch_id = upload(client, admin_token, content, import_type="REGULAR_PRODUCT").json()["id"]
    for row_number, raw_value in enumerate(("=1+1", "+SUM(A1:A2)", "-CMD", "@IMPORT", "-12.5"), start=2):
        db.add(ImportRowIssue(import_batch_id=batch_id, sheet_name="数据", excel_row_number=row_number, severity="WARNING", field_name="测试字段", raw_value=raw_value, issue_code="CSV_TEST", message="中文测试"))
    db.commit()
    response = client.get(f"/api/v1/imports/{batch_id}/issues/export", headers={"Authorization": f"Bearer {admin_token}"})
    text = response.content.decode("utf-8-sig")
    assert response.content.startswith(b"\xef\xbb\xbf")
    assert "'=1+1" in text
    assert "'+SUM(A1:A2)" in text
    assert "'-CMD" in text
    assert "'@IMPORT" in text
    assert "-12.5" in text and "'-12.5" not in text
    assert "中文测试" in text
