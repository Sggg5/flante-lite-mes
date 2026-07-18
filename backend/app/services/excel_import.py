from __future__ import annotations

import hashlib
import json
import re
import zipfile
from collections import Counter
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable, Iterator

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.datetime import from_excel
from xml.etree.ElementTree import iterparse
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from app.models import (
    FittingWipSnapshot,
    ImportBatch,
    ImportedWeeklyPlanRaw,
    ImportRowIssue,
    InventorySnapshot,
    PipeWipSnapshot,
    Product,
    ProductImportChange,
    RegularProductionProduct,
    ShipmentRecord,
    WeeklyPlanStagingRow,
)


IMPORT_TYPES = {
    "SHIPMENT",
    "INVENTORY",
    "PIPE_WIP",
    "FITTING_WIP",
    "REGULAR_PRODUCT",
    "WEEKLY_PLAN",
}
EMPTY_ROW_STOP = 200
MAX_SCAN_ROWS = 500_000
SCIENCE_PATTERN = re.compile(r"^[+-]?\d+(?:\.\d+)?[Ee][+-]?\d+$")

FIELD_SYNONYMS: dict[str, tuple[str, ...]] = {
    "product_code": ("存货编码", "产品编码", "物料编码", "编码", "代码"),
    "product_name": ("存货名称", "产品名称", "名称", "品名"),
    "specification": ("规格型号", "规格", "型号"),
    "category": ("存货分类", "产品分类", "分类", "类别"),
    "unit": ("主计量单位", "计量单位", "单位"),
    "quantity": ("出库数量", "未入库数量", "未完成数量", "未完成", "数量"),
    "document_no": ("出库单号", "单据号", "单号"),
    "shipment_date": ("出库日期", "日期"),
    "shipment_month": ("出库月份", "月份"),
    "production_batch_no": ("生产批次号", "生产批次", "批次号", "批次", "批号"),
    "snapshot_date": ("快照日期", "统计日期", "数据日期"),
    "on_hand_qty": ("现存数量", "现存", "库存数量"),
    "expected_inbound_qty": ("预计入库数量", "预计入库"),
    "expected_outbound_qty": ("预计出库数量", "预计出库"),
    "source_available_qty": ("可用数量", "可用量"),
    "process_name": ("工序名称", "工序"),
    "equipment_name": ("设备名称", "设备", "机台"),
    "planned_quantity": ("周计划", "计划数量", "计划量"),
    "actual_quantity": ("周实际", "实际数量", "实际量"),
    "plan_start_date": ("计划开始日期", "开始日期"),
    "plan_end_date": ("计划结束日期", "结束日期"),
    "row_kind": ("计划/实际", "类型", "行类型"),
}

TYPE_FIELDS: dict[str, tuple[str, ...]] = {
    "SHIPMENT": ("product_code", "product_name", "specification", "unit", "document_no", "shipment_date", "shipment_month", "quantity", "production_batch_no"),
    "INVENTORY": ("product_code", "product_name", "specification", "category", "unit", "snapshot_date", "on_hand_qty", "expected_inbound_qty", "expected_outbound_qty", "source_available_qty"),
    "PIPE_WIP": ("product_code", "product_name", "specification", "snapshot_date", "quantity"),
    "FITTING_WIP": ("product_code", "product_name", "specification", "snapshot_date", "production_batch_no", "quantity"),
    "REGULAR_PRODUCT": ("product_code", "product_name", "specification", "category", "unit"),
    "WEEKLY_PLAN": ("product_code", "product_name", "specification", "production_batch_no", "process_name", "equipment_name", "row_kind", "plan_start_date", "plan_end_date", "planned_quantity", "actual_quantity"),
}

REQUIRED_FIELDS: dict[str, set[str]] = {
    "SHIPMENT": {"product_code", "document_no", "shipment_date", "quantity"},
    "INVENTORY": {"product_code", "on_hand_qty", "expected_inbound_qty", "expected_outbound_qty"},
    "PIPE_WIP": {"product_code", "quantity"},
    "FITTING_WIP": {"product_code", "quantity"},
    "REGULAR_PRODUCT": {"product_code"},
    "WEEKLY_PLAN": {"product_code", "production_batch_no", "process_name", "equipment_name", "planned_quantity"},
}

WEEKLY_SHEET_LAYOUTS: dict[str, dict[str, Any]] = {
    "制管": {"header_end": 5, "equipment": 1, "name": 3, "spec": 4, "batch": 5, "process": 6, "row_kind": 12, "total": 13, "daily": range(14, 21)},
    "包装": {"header_end": 4, "equipment": 1, "name": 2, "spec": 3, "batch": 4, "process": 5, "row_kind": 10, "total": 11, "daily": range(12, 19)},
    "成型": {"header_end": 5, "equipment": 1, "name": 3, "spec": 4, "batch": 5, "process": 6, "row_kind": 11, "total": 12, "daily": range(13, 20)},
    "下料": {"header_end": 4, "equipment": None, "name": 1, "spec": 2, "batch": 3, "process": None, "row_kind": 8, "total": 9, "daily": range(10, 17)},
}


class ImportValidationError(ValueError):
    def __init__(self, code: str, message: str, details: Any = None):
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = details


def safe_filename(filename: str | None) -> str:
    name = Path(filename or "upload.xlsx").name
    cleaned = re.sub(r"[^\w.\-()\u4e00-\u9fff ]", "_", name, flags=re.UNICODE).strip(" .")
    return (cleaned or "upload.xlsx")[:255]


def sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def make_batch_no() -> str:
    return f"IMP-{datetime.now(UTC):%Y%m%d%H%M%S}-{hashlib.sha1(str(datetime.now(UTC).timestamp()).encode()).hexdigest()[:8].upper()}"


def find_duplicate_import_batch(
    db: Session,
    *,
    import_type: str,
    file_sha256: str,
    source_date: date | None,
    exclude_batch_id: int | None = None,
    lock: bool = False,
) -> ImportBatch | None:
    if lock and db.get_bind().dialect.name == "postgresql":
        identity = f"{import_type}|{file_sha256}|{source_date.isoformat() if source_date else '<NULL>'}"
        advisory_key = int(hashlib.sha256(identity.encode("utf-8")).hexdigest()[:16], 16) & ((1 << 63) - 1)
        db.execute(select(func.pg_advisory_xact_lock(advisory_key)))
    statement = select(ImportBatch).where(
        ImportBatch.import_type == import_type,
        ImportBatch.file_sha256 == file_sha256,
        ImportBatch.status.notin_({"CANCELLED", "ROLLED_BACK"}),
    )
    statement = statement.where(
        ImportBatch.source_date.is_(None) if source_date is None else ImportBatch.source_date == source_date
    )
    if exclude_batch_id is not None:
        statement = statement.where(ImportBatch.id != exclude_batch_id)
    if lock:
        statement = statement.with_for_update()
    return db.scalar(statement.order_by(ImportBatch.id).limit(1))


def load_safe_workbook(path: Path, *, read_only: bool = True):
    try:
        return load_workbook(path, read_only=read_only, data_only=False, keep_links=False)
    except Exception as exc:
        raise ImportValidationError("INVALID_WORKBOOK", "Excel 工作簿损坏、加密或无法解析") from exc


def row_has_value(row: Iterable[Cell]) -> bool:
    return any(cell.value not in (None, "") for cell in row)


def scan_effective_range(worksheet) -> dict[str, Any]:
    last_row = 0
    last_column = 0
    empty_run = 0
    first_row = 0
    scanned_rows = 0
    formula_count = 0
    external_formula_count = 0
    error_cell_count = 0
    for row_number, row in enumerate(worksheet.iter_rows(), start=1):
        scanned_rows = row_number
        nonempty_columns = [cell.column for cell in row if cell.value not in (None, "")]
        if nonempty_columns:
            first_row = first_row or row_number
            last_row = row_number
            last_column = max(last_column, max(nonempty_columns))
            empty_run = 0
            for cell in row:
                if cell.data_type == "f":
                    formula_count += 1
                    if "[" in str(cell.value):
                        external_formula_count += 1
                elif cell.data_type == "e":
                    error_cell_count += 1
        elif first_row:
            empty_run += 1
            if empty_run >= EMPTY_ROW_STOP:
                break
        if row_number >= MAX_SCAN_ROWS:
            break
    return {
        "first_row": first_row,
        "last_row": last_row,
        "last_column": last_column,
        "scanned_rows": scanned_rows,
        "formula_count": formula_count,
        "external_formula_count": external_formula_count,
        "error_cell_count": error_cell_count,
        "scan_truncated": scanned_rows >= MAX_SCAN_ROWS,
    }


def extract_sheet_structure(path: Path, worksheet) -> dict[str, Any]:
    hidden_rows: set[int] = set()
    hidden_columns: set[int] = set()
    merged_ranges: list[str] = []
    auto_filter_range: str | None = None
    worksheet_path = getattr(worksheet, "_worksheet_path", None)
    if not worksheet_path:
        return {"hidden_rows": hidden_rows, "hidden_columns": hidden_columns, "merged_ranges": merged_ranges, "auto_filter_range": auto_filter_range}
    with zipfile.ZipFile(path) as archive, archive.open(worksheet_path.lstrip("/")) as source:
        for _, element in iterparse(source, events=("end",)):
            tag = element.tag.rsplit("}", 1)[-1]
            if tag == "row" and element.attrib.get("hidden") in {"1", "true"}:
                hidden_rows.add(int(element.attrib["r"]))
            elif tag == "col" and element.attrib.get("hidden") in {"1", "true"}:
                hidden_columns.update(range(int(element.attrib["min"]), int(element.attrib["max"]) + 1))
            elif tag == "mergeCell" and element.attrib.get("ref"):
                merged_ranges.append(element.attrib["ref"])
            elif tag == "autoFilter":
                auto_filter_range = element.attrib.get("ref")
            element.clear()
    return {"hidden_rows": hidden_rows, "hidden_columns": hidden_columns, "merged_ranges": merged_ranges, "auto_filter_range": auto_filter_range}


def merged_cell_lookup(ranges: list[str]) -> dict[tuple[int, int], tuple[int, int]]:
    lookup: dict[tuple[int, int], tuple[int, int]] = {}
    for cell_range in ranges:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        if (max_row - min_row + 1) * (max_col - min_col + 1) > 10_000:
            continue
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                lookup[(row, column)] = (min_row, min_col)
    return lookup


def normalize_header(value: Any) -> str:
    return re.sub(r"[\s\n\r　]+", "", str(value or "")).lower()


def detect_header(rows: list[list[Any]], import_type: str) -> tuple[int, int, dict[str, int]]:
    fields = TYPE_FIELDS[import_type]
    best_end = 1
    best_score = -1
    for row_index, row in enumerate(rows[:12], start=1):
        normalized = [normalize_header(value) for value in row]
        score = sum(
            1
            for field in fields
            if any(normalize_header(alias) in header for alias in FIELD_SYNONYMS[field] for header in normalized if header)
        )
        if score > best_score:
            best_score = score
            best_end = row_index
    start = best_end
    if best_end > 1 and any(value not in (None, "") for value in rows[best_end - 2]):
        start = max(1, best_end - 2)
    combined_headers: dict[int, str] = {}
    width = max((len(row) for row in rows[start - 1 : best_end]), default=0)
    for column in range(width):
        parts: list[str] = []
        for row in rows[start - 1 : best_end]:
            if column < len(row) and row[column] not in (None, ""):
                value = str(row[column]).strip()
                if not parts or parts[-1] != value:
                    parts.append(value)
        combined_headers[column + 1] = " ".join(parts)
    return start, best_end, auto_mapping(combined_headers, import_type)


def auto_mapping(headers: dict[int, str], import_type: str) -> dict[str, int]:
    mapping: dict[str, int] = {}
    used_columns: set[int] = set()
    for field in TYPE_FIELDS[import_type]:
        aliases = tuple(normalize_header(item) for item in FIELD_SYNONYMS[field])
        for column, header in headers.items():
            normalized = normalize_header(header)
            if column not in used_columns and any(alias == normalized or alias in normalized for alias in aliases):
                mapping[field] = column
                used_columns.add(column)
                break
    return mapping


def analyze_workbook(path: Path, import_type: str, selected_sheet: str | None = None) -> dict[str, Any]:
    workbook = load_safe_workbook(path)
    try:
        if selected_sheet and selected_sheet not in workbook.sheetnames:
            raise ImportValidationError("SHEET_NOT_FOUND", "指定工作表不存在")
        sheet_results: list[dict[str, Any]] = []
        target_names = [selected_sheet] if selected_sheet else workbook.sheetnames
        workbook_year: int | None = None
        if import_type == "WEEKLY_PLAN":
            for workbook_sheet_name in workbook.sheetnames:
                candidate_layout = WEEKLY_SHEET_LAYOUTS.get(workbook_sheet_name)
                if candidate_layout is None:
                    continue
                candidate_sheet = workbook[workbook_sheet_name]
                title_text = " ".join(
                    str(cell.value)
                    for row in candidate_sheet.iter_rows(min_row=1, max_row=candidate_layout["header_end"])
                    for cell in row
                    if cell.value not in (None, "")
                )
                if year_match := re.search(r"(20\d{2})", title_text):
                    workbook_year = int(year_match.group(1))
                    break
        for name in target_names:
            worksheet = workbook[name]
            effective = scan_effective_range(worksheet)
            structure = extract_sheet_structure(path, worksheet)
            preview_rows: list[list[Any]] = []
            for row in worksheet.iter_rows(min_row=1, max_row=min(max(effective["last_row"], 1), 12)):
                preview_rows.append([cell.value for cell in row[: max(effective["last_column"], 1)]])
            header_start, header_end, mapping = detect_header(preview_rows, import_type)
            weekly_dates = infer_week_dates(worksheet, WEEKLY_SHEET_LAYOUTS[name], workbook_year) if import_type == "WEEKLY_PLAN" and name in WEEKLY_SHEET_LAYOUTS else None
            detected_dates = weekly_dates or sorted({
                    parsed
                    for row in preview_rows[:header_end]
                    for value in row
                    if isinstance(value, (date, datetime)) or (isinstance(value, (int, float)) and value > 30_000) or (isinstance(value, str) and re.search(r"20\d{2}|\d{1,2}[./-]\d{1,2}", value))
                    if (parsed := parse_date(value)) is not None
                })
            sheet_results.append(
                {
                    "sheet_name": name,
                    "declared_rows": worksheet.max_row,
                    "declared_columns": worksheet.max_column,
                    **effective,
                    "header_row_start": header_start,
                    "header_row_end": header_end,
                    "auto_mapping": mapping,
                    "preview_headers": preview_rows[:header_end],
                    "hidden_row_count": len(structure["hidden_rows"]),
                    "hidden_column_count": len(structure["hidden_columns"]),
                    "merged_ranges": structure["merged_ranges"],
                    "auto_filter_range": structure["auto_filter_range"],
                    "detected_date_start": detected_dates[0].isoformat() if detected_dates else None,
                    "detected_date_end": detected_dates[-1].isoformat() if detected_dates else None,
                }
            )
        periods = {(item["detected_date_start"], item["detected_date_end"]) for item in sheet_results if item["detected_date_start"]}
        return {"sheet_count": len(workbook.sheetnames), "sheet_names": workbook.sheetnames, "sheets": sheet_results, "plan_period_consistent": len(periods) <= 1}
    finally:
        workbook.close()


def normalize_product_code(cell: Cell) -> tuple[str | None, str | None]:
    value = cell.value
    if value in (None, ""):
        return None, "PRODUCT_CODE_REQUIRED"
    if isinstance(value, str):
        text = value.strip()
        if SCIENCE_PATTERN.fullmatch(text):
            return text, "PRODUCT_CODE_SCIENTIFIC_NOTATION"
        return text, None
    if isinstance(value, bool):
        return str(value), "PRODUCT_CODE_FORMAT_INVALID"
    if isinstance(value, (int, float, Decimal)):
        number_format = str(cell.number_format or "")
        zero_mask = re.fullmatch(r"0+", number_format)
        if zero_mask and float(value).is_integer():
            return str(int(value)).zfill(len(number_format)), None
        if isinstance(value, float) and not value.is_integer():
            return format(value, "f").rstrip("0").rstrip("."), "PRODUCT_CODE_NUMERIC_AMBIGUOUS"
        return str(int(value)), "PRODUCT_CODE_NUMERIC_AMBIGUOUS"
    return str(value).strip(), "PRODUCT_CODE_FORMAT_INVALID"


def parse_decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None


def parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except (ValueError, OverflowError):
            return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日", "%Y-%m", "%Y/%m", "%Y年%m月"):
        try:
            parsed = datetime.strptime(text, fmt).date()
            return parsed.replace(day=1) if "%d" not in fmt else parsed
        except ValueError:
            continue
    return None


def infer_week_dates(worksheet, layout: dict[str, Any], default_year: int | None = None) -> list[date] | None:
    title_values = [
        str(cell.value)
        for row in worksheet.iter_rows(min_row=1, max_row=layout["header_end"], values_only=False)
        for cell in row
        if cell.value not in (None, "")
    ]
    title = " ".join(title_values)
    full_dates = re.findall(r"(20\d{2})\D{0,3}(\d{1,2})\D{0,3}(\d{1,2})", title)
    start: date | None = None
    if full_dates:
        year, month, day = map(int, full_dates[0])
        try:
            start = date(year, month, day)
        except ValueError:
            start = None
    if start is None:
        year_match = re.search(r"(20\d{2})", title)
        range_match = re.search(r"(\d{1,2})[./-](\d{1,2})\s*(?:至|到|[-~—])\s*(\d{1,2})[./-](\d{1,2})", title)
        inferred_year = int(year_match.group(1)) if year_match else default_year
        if inferred_year and range_match:
            try:
                start = date(inferred_year, int(range_match.group(1)), int(range_match.group(2)))
            except ValueError:
                start = None
    header_dates: list[date] = []
    for column in layout["daily"]:
        parsed: date | None = None
        day_number: int | None = None
        for row_number in range(1, layout["header_end"] + 1):
            value = worksheet.cell(row_number, column).value
            if isinstance(value, (date, datetime)):
                parsed = value.date() if isinstance(value, datetime) else value
            elif isinstance(value, (int, float)):
                if 1 <= int(value) <= 31:
                    day_number = int(value)
                else:
                    parsed = parse_date(value)
            elif isinstance(value, str):
                parsed = parse_date(value)
                day_match = re.fullmatch(r"\s*(\d{1,2})\s*(?:日|号)?\s*", value)
                if day_match:
                    day_number = int(day_match.group(1))
            if parsed:
                break
        if parsed:
            header_dates.append(parsed)
        elif start and day_number:
            month = start.month + (1 if day_number < start.day else 0)
            year = start.year + (1 if month == 13 else 0)
            month = 1 if month == 13 else month
            try:
                header_dates.append(date(year, month, day_number))
            except ValueError:
                return None
        elif start:
            header_dates.append(start.fromordinal(start.toordinal() + len(header_dates)))
        else:
            return None
    return header_dates if len(header_dates) == 7 else None


def iter_weekly_plan_rows(path: Path, batch: ImportBatch) -> Iterator[tuple[int, dict[str, Any], list[dict[str, Any]]]]:
    sheet_name = batch.selected_sheet_name or ""
    layout = WEEKLY_SHEET_LAYOUTS.get(sheet_name)
    if layout is None:
        raise ImportValidationError("WEEKLY_PLAN_SHEET_UNSUPPORTED", "周计划工作表必须为制管、包装、成型或下料")
    options = batch.import_options or {}
    include_hidden_rows = bool(options.get("include_hidden_rows", True))
    workbook = load_safe_workbook(path)
    try:
        worksheet = workbook[sheet_name]
        structure = extract_sheet_structure(path, worksheet)
        merged_lookup = merged_cell_lookup(structure["merged_ranges"])
        hidden_rows = structure["hidden_rows"]
        stored_period = (options.get("sheet_periods") or {}).get(sheet_name) or []
        stored_start = parse_date(stored_period[0]) if len(stored_period) > 0 else None
        stored_end = parse_date(stored_period[1]) if len(stored_period) > 1 else None
        if stored_start and stored_end and (stored_end - stored_start).days == 6:
            week_dates = [stored_start.fromordinal(stored_start.toordinal() + offset) for offset in range(7)]
        else:
            default_year = batch.source_date.year if batch.source_date else None
            week_dates = infer_week_dates(worksheet, layout, default_year)
        master_cells: dict[tuple[int, int], Cell] = {}

        def resolved_cell(row: tuple[Cell, ...], row_number: int, column: int | None) -> Cell | None:
            if not column or column > len(row):
                return None
            cell = row[column - 1]
            if cell.value in (None, ""):
                master = merged_lookup.get((row_number, column))
                if master:
                    return master_cells.get(master, cell)
            return cell

        rows = list(worksheet.iter_rows(min_row=layout["header_end"] + 1))
        for offset, row in enumerate(rows):
            row_number = layout["header_end"] + 1 + offset
            for column_number, cell in enumerate(row, start=1):
                coordinate = (row_number, column_number)
                if cell.value not in (None, "") and merged_lookup.get(coordinate, coordinate) == coordinate:
                    master_cells[coordinate] = cell
            row_kind_cell = resolved_cell(row, row_number, layout["row_kind"])
            row_kind = str(row_kind_cell.value or "").strip()
            if "计划" not in row_kind:
                continue
            source_row_hidden = row_number in hidden_rows
            if source_row_hidden and not include_hidden_rows:
                yield row_number, {}, [issue(row_number, "INFO", None, None, "HIDDEN_ROW_EXCLUDED", "用户选择仅导入可见行，该隐藏计划行已排除")]
                continue
            actual_row = rows[offset + 1] if offset + 1 < len(rows) else tuple()
            actual_row_number = row_number + 1
            actual_kind_cell = resolved_cell(actual_row, actual_row_number, layout["row_kind"]) if actual_row else None
            has_actual_row = "实际" in str(actual_kind_cell.value or "")
            actual_row_hidden = has_actual_row and actual_row_number in hidden_rows
            actual_values_excluded = bool(actual_row_hidden and not include_hidden_rows)
            problems: list[dict[str, Any]] = []
            if not has_actual_row:
                problems.append(issue(row_number, "WARNING", "row_kind", None, "WEEKLY_PLAN_ACTUAL_ROW_MISSING", "未找到配对实际行，计划仍进入待匹配区"))
            elif actual_values_excluded:
                problems.append(issue(row_number, "INFO", "daily_actual", None, "WEEKLY_ACTUAL_HIDDEN_EXCLUDED", "实际行已隐藏且用户选择仅导入可见行，实际量未读取"))
            if week_dates is None:
                problems.append(issue(row_number, "ERROR", "plan_start_date", None, "WEEKLY_PLAN_PERIOD_UNRECOGNIZED", "无法从标题和日期表头识别七天计划周期"))
            name = resolved_cell(row, row_number, layout["name"])
            spec = resolved_cell(row, row_number, layout["spec"])
            batch_cell = resolved_cell(row, row_number, layout["batch"])
            process_cell = resolved_cell(row, row_number, layout["process"])
            equipment_cell = resolved_cell(row, row_number, layout["equipment"])
            production_batch_no = str(batch_cell.value).strip() if batch_cell and batch_cell.value not in (None, "") else None
            process_name = str(process_cell.value).strip() if process_cell and process_cell.value not in (None, "") else sheet_name
            equipment_name = str(equipment_cell.value).strip() if equipment_cell and equipment_cell.value not in (None, "") else sheet_name if sheet_name == "下料" else None
            if not production_batch_no:
                problems.append(issue(row_number, "ERROR", "production_batch_no", None, "WEEKLY_PLAN_BATCH_REQUIRED", "生产批次号缺失"))
            if not equipment_name:
                problems.append(issue(row_number, "WARNING", "equipment_name", None, "WEEKLY_PLAN_EQUIPMENT_MISSING", "设备为空，需在待匹配区确认"))
            daily_plan: dict[str, str | None] = {}
            daily_actual: dict[str, str | None] = {}
            formula_metadata: dict[str, Any] = {"plan_formulas": {}, "actual_formulas": {}, "external_actual_formulas": {}}
            for index, column in enumerate(layout["daily"]):
                date_key = week_dates[index].isoformat() if week_dates else f"DAY_{index + 1}"
                plan_cell = resolved_cell(row, row_number, column)
                actual_cell = resolved_cell(actual_row, actual_row_number, column) if has_actual_row and not actual_values_excluded else None
                plan_value = parse_decimal(plan_cell.value) if plan_cell and plan_cell.data_type != "f" else None
                actual_value = parse_decimal(actual_cell.value) if actual_cell and actual_cell.data_type != "f" else None
                daily_plan[date_key] = str(plan_value) if plan_value is not None else None
                daily_actual[date_key] = str(actual_value) if actual_value is not None else None
                if plan_cell and plan_cell.data_type == "f":
                    formula_metadata["plan_formulas"][date_key] = str(plan_cell.value)
                if actual_cell and actual_cell.data_type == "f":
                    formula = str(actual_cell.value)
                    if "[" in formula:
                        formula_metadata["external_actual_formulas"][date_key] = formula
                    else:
                        formula_metadata["actual_formulas"][date_key] = formula
            if formula_metadata["external_actual_formulas"]:
                problems.append(issue(row_number, "WARNING", "daily_actual", None, "WEEKLY_ACTUAL_EXTERNAL_FORMULA", "实际量来自外部日报公式，未执行；计划量仍可进入待匹配区"))
            weekly_plan = sum((Decimal(value) for value in daily_plan.values() if value is not None), Decimal("0"))
            weekly_actual_values = [Decimal(value) for value in daily_actual.values() if value is not None]
            weekly_actual = sum(weekly_actual_values, Decimal("0")) if weekly_actual_values else None
            if weekly_plan == 0:
                total_cell = resolved_cell(row, row_number, layout["total"])
                if total_cell and total_cell.data_type != "f":
                    weekly_plan = parse_decimal(total_cell.value) or Decimal("0")
                elif total_cell and total_cell.data_type == "f":
                    formula_metadata["weekly_plan_formula"] = str(total_cell.value)
            if weekly_plan == 0:
                problems.append(issue(row_number, "WARNING", "weekly_plan_qty", 0, "WEEKLY_PLAN_ZERO", "周计划数量为0，请人工确认"))
            normalized = {
                "_weekly_staging": True,
                "product_name": str(name.value).strip() if name and name.value not in (None, "") else None,
                "specification": str(spec.value).strip() if spec and spec.value not in (None, "") else None,
                "production_batch_no": production_batch_no,
                "process_name": process_name,
                "equipment_name": equipment_name,
                "plan_start_date": week_dates[0] if week_dates else None,
                "plan_end_date": week_dates[-1] if week_dates else None,
                "daily_plan": daily_plan,
                "daily_actual": daily_actual,
                "planned_quantity": weekly_plan,
                "actual_quantity": weekly_actual,
                "formula_metadata": formula_metadata,
            }
            normalized["raw_data"] = {
                "source_row_hidden": source_row_hidden,
                "actual_row_hidden": actual_row_hidden,
                "actual_row_excluded": actual_values_excluded,
                "plan_row_number": row_number,
                "actual_row_number": actual_row_number if has_actual_row else None,
                "product_name_raw": normalized["product_name"],
                "specification_raw": normalized["specification"],
                "production_batch_no": production_batch_no,
                "process_name": process_name,
                "equipment_name": equipment_name,
                "daily_plan": daily_plan,
                "daily_actual": daily_actual,
                "formula_metadata": formula_metadata,
            }
            yield row_number, normalized, problems
    finally:
        workbook.close()


def issue(row: int, severity: str, field: str | None, raw: Any, code: str, message: str) -> dict[str, Any]:
    return {
        "excel_row_number": row,
        "severity": severity,
        "field_name": field,
        "raw_value": None if raw is None else str(raw)[:2000],
        "issue_code": code,
        "message": message,
    }


def mapped_cell(row: tuple[Cell, ...], mapping: dict[str, int], field: str) -> Cell | None:
    column = mapping.get(field)
    return row[column - 1] if column and column <= len(row) else None


def iter_normalized_rows(
    path: Path,
    batch: ImportBatch,
) -> Iterator[tuple[int, dict[str, Any], list[dict[str, Any]]]]:
    options = batch.import_options or {}
    sheet_name = batch.selected_sheet_name
    if not sheet_name:
        raise ImportValidationError("SHEET_REQUIRED", "请先选择工作表并完成分析")
    if batch.import_type == "WEEKLY_PLAN":
        yield from iter_weekly_plan_rows(path, batch)
        return
    mapping = {key: int(value) for key, value in (batch.field_mapping or {}).items()}
    missing = sorted(REQUIRED_FIELDS[batch.import_type] - mapping.keys())
    if missing:
        raise ImportValidationError("MAPPING_REQUIRED", "缺少必填字段映射", {"fields": missing})
    if len(set(mapping.values())) != len(mapping.values()):
        raise ImportValidationError("DUPLICATE_FIELD_MAPPING", "一个 Excel 列不能映射到多个业务字段")
    workbook = load_safe_workbook(path)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ImportValidationError("SHEET_NOT_FOUND", "指定工作表不存在")
        worksheet = workbook[sheet_name]
        structure = extract_sheet_structure(path, worksheet)
        hidden_rows = structure["hidden_rows"]
        include_hidden_rows = bool(options.get("include_hidden_rows", True))
        merged_lookup = merged_cell_lookup(structure["merged_ranges"])
        master_cells: dict[tuple[int, int], Cell] = {}
        header_end = int(options.get("header_row_end", 1))
        empty_run = 0
        seen_data = False
        for row_number, row in enumerate(worksheet.iter_rows(min_row=header_end + 1), start=header_end + 1):
            for column_number, cell in enumerate(row, start=1):
                coordinate = (row_number, column_number)
                if cell.value not in (None, "") and merged_lookup.get(coordinate, coordinate) == coordinate:
                    master_cells[coordinate] = cell
            source_row_hidden = row_number in hidden_rows
            if source_row_hidden and not include_hidden_rows:
                if row_has_value(row):
                    yield row_number, {}, [issue(row_number, "INFO", None, None, "HIDDEN_ROW_EXCLUDED", "用户选择仅导入可见行，该隐藏数据行已排除")]
                continue
            if not row_has_value(row):
                if seen_data:
                    empty_run += 1
                    if empty_run >= EMPTY_ROW_STOP:
                        break
                    yield row_number, {}, [issue(row_number, "INFO", None, None, "BLANK_ROW_SKIPPED", "空白数据行已跳过")]
                continue
            seen_data = True
            empty_run = 0
            normalized: dict[str, Any] = {}
            problems: list[dict[str, Any]] = []
            for field in TYPE_FIELDS[batch.import_type]:
                cell = mapped_cell(row, mapping, field)
                column = mapping.get(field)
                if cell and cell.value in (None, "") and column:
                    master = merged_lookup.get((row_number, column))
                    if master:
                        cell = master_cells.get(master, cell)
                value = cell.value if cell else None
                if cell and cell.data_type == "e":
                    problems.append(issue(row_number, "ERROR", field, value, "EXCEL_ERROR_CELL", "Excel 错误单元格无法导入"))
                if cell and cell.data_type == "f":
                    severity = "ERROR" if "[" in str(value) else "WARNING"
                    code = "EXTERNAL_FORMULA_REFERENCE" if severity == "ERROR" else "FORMULA_VALUE_UNTRUSTED"
                    problems.append(issue(row_number, severity, field, value, code, "公式结果不作为可信业务值"))
                if field == "product_code" and cell:
                    code, code_problem = normalize_product_code(cell)
                    normalized[field] = code
                    if code_problem:
                        severity = "ERROR" if code_problem == "PRODUCT_CODE_SCIENTIFIC_NOTATION" else "WARNING"
                        problems.append(issue(row_number, severity, field, value, code_problem, "产品编码格式需要确认"))
                elif field.endswith("_date") or field == "shipment_month":
                    normalized[field] = parse_date(value)
                    if value not in (None, "") and normalized[field] is None:
                        problems.append(issue(row_number, "ERROR", field, value, "DATE_INVALID", "日期无法转换"))
                elif field in {"quantity", "on_hand_qty", "expected_inbound_qty", "expected_outbound_qty", "source_available_qty", "planned_quantity", "actual_quantity"}:
                    normalized[field] = parse_decimal(value)
                    weekly_actual = batch.import_type == "WEEKLY_PLAN" and "实际" in str(normalized.get("row_kind") or "")
                    if field in REQUIRED_FIELDS[batch.import_type] and normalized[field] is None and not (field == "planned_quantity" and weekly_actual):
                        problems.append(issue(row_number, "ERROR", field, value, "QUANTITY_INVALID", "数量为空或无法转换"))
                else:
                    normalized[field] = str(value).strip() if value not in (None, "") else None
            for field in REQUIRED_FIELDS[batch.import_type]:
                weekly_actual = batch.import_type == "WEEKLY_PLAN" and "实际" in str(normalized.get("row_kind") or "")
                if normalized.get(field) in (None, "") and not (field == "planned_quantity" and weekly_actual):
                    problems.append(issue(row_number, "ERROR", field, None, "REQUIRED_FIELD_MISSING", "必填字段为空"))
            if batch.import_type == "WEEKLY_PLAN" and "实际" in str(normalized.get("row_kind") or ""):
                if normalized.get("actual_quantity") is None:
                    normalized["actual_quantity"] = normalized.get("planned_quantity")
                normalized["planned_quantity"] = None
            quantity_field = "quantity" if "quantity" in normalized else "planned_quantity"
            quantity = normalized.get(quantity_field)
            if batch.import_type in {"SHIPMENT", "PIPE_WIP", "FITTING_WIP"} and quantity is not None and quantity < 0:
                problems.append(issue(row_number, "WARNING", quantity_field, quantity, "NEGATIVE_QUANTITY", "负数量保留原值，请人工确认"))
            if batch.import_type == "INVENTORY" and all(normalized.get(name) is not None for name in ("on_hand_qty", "expected_inbound_qty", "expected_outbound_qty")):
                calculated = normalized["on_hand_qty"] + normalized["expected_inbound_qty"] - normalized["expected_outbound_qty"]
                normalized["calculated_available_qty"] = calculated
                source = normalized.get("source_available_qty")
                if source is not None and source != calculated:
                    problems.append(issue(row_number, "WARNING", "source_available_qty", source, "AVAILABLE_QTY_MISMATCH", "Excel 可用量与系统复算值不一致"))
            if batch.import_type in {"INVENTORY", "PIPE_WIP", "FITTING_WIP"}:
                effective_snapshot_date = normalized.get("snapshot_date") or parse_date(options.get("source_date"))
                normalized["snapshot_date"] = effective_snapshot_date
                if effective_snapshot_date is None:
                    problems.append(issue(row_number, "ERROR", "snapshot_date", None, "SNAPSHOT_DATE_REQUIRED", "Excel 快照日期和批次数据日期均为空"))
            if batch.import_type == "WEEKLY_PLAN" and normalized.get("plan_start_date") is None and normalized.get("plan_end_date") is None:
                problems.append(issue(row_number, "ERROR", "plan_start_date", None, "WEEKLY_PLAN_PERIOD_UNRECOGNIZED", "计划周期无法识别"))
            normalized["raw_data"] = {
                field: (value.isoformat() if isinstance(value, (date, datetime)) else str(value) if isinstance(value, Decimal) else value)
                for field, value in normalized.items()
                if field != "raw_data"
            }
            normalized["raw_data"]["source_row_hidden"] = source_row_hidden
            yield row_number, normalized, problems
            if row_number >= MAX_SCAN_ROWS:
                break
    finally:
        workbook.close()


def validate_batch(db: Session, batch: ImportBatch, path: Path) -> dict[str, Any]:
    db.execute(delete(ImportRowIssue).where(ImportRowIssue.import_batch_id == batch.id))
    total = 0
    valid = 0
    warning_rows: set[int] = set()
    error_rows: set[int] = set()
    seen_keys: set[tuple[Any, ...]] = set()
    product_identity: dict[str, tuple[str | None, str | None]] = {}
    existing_products: dict[str, Product | None] = {}
    hidden_data_rows = 0
    excluded_hidden_data_rows = 0
    issue_counts: Counter[str] = Counter()
    for row_number, normalized, problems in iter_normalized_rows(path, batch):
        if not normalized:
            for problem in problems:
                if problem["issue_code"] == "HIDDEN_ROW_EXCLUDED":
                    hidden_data_rows += 1
                    excluded_hidden_data_rows += 1
                issue_counts[problem["issue_code"]] += 1
                db.add(ImportRowIssue(import_batch_id=batch.id, sheet_name=batch.selected_sheet_name or "", **problem))
            continue
        total += 1
        raw_data = normalized.get("raw_data", {})
        if raw_data.get("source_row_hidden"):
            hidden_data_rows += 1
        if raw_data.get("actual_row_hidden"):
            hidden_data_rows += 1
        if raw_data.get("actual_row_excluded"):
            excluded_hidden_data_rows += 1
        code = normalized.get("product_code")
        identity = (normalized.get("product_name"), normalized.get("specification"))
        if code and code in product_identity and product_identity[code] != identity:
            problems.append(issue(row_number, "WARNING", "product_code", code, "PRODUCT_IDENTITY_CONFLICT", "同一编码的名称或规格不一致"))
        elif code:
            product_identity[code] = identity
        if code:
            if code not in existing_products:
                existing_products[code] = db.scalar(select(Product).where(Product.product_code == code))
            existing_product = existing_products[code]
            if existing_product:
                conflict_fields = {
                    "product_name": ("PRODUCT_MASTER_NAME_CONFLICT", existing_product.product_name),
                    "specification": ("PRODUCT_MASTER_SPEC_CONFLICT", existing_product.specification),
                    "category": ("PRODUCT_MASTER_CATEGORY_CONFLICT", existing_product.category),
                    "unit": ("PRODUCT_MASTER_UNIT_CONFLICT", existing_product.unit),
                }
                for field, (conflict_code, database_value) in conflict_fields.items():
                    excel_value = normalized.get(field)
                    if excel_value not in (None, "") and database_value not in (None, "") and excel_value != database_value:
                        raw_comparison = json.dumps({"excel_value": excel_value, "database_value": database_value}, ensure_ascii=False)
                        problems.append(issue(row_number, "WARNING", field, raw_comparison, conflict_code, "导入值与现有产品主数据不一致，默认不会覆盖"))
        if batch.import_type == "SHIPMENT":
            key = (normalized.get("document_no"), code, normalized.get("shipment_date"), normalized.get("quantity"))
        elif batch.import_type == "FITTING_WIP":
            key = (code, normalized.get("production_batch_no"), normalized.get("quantity"))
        elif batch.import_type == "WEEKLY_PLAN":
            key = (normalized.get("product_name"), normalized.get("specification"), normalized.get("production_batch_no"), normalized.get("process_name"), row_number)
        else:
            key = (code,)
        if key in seen_keys:
            problems.append(issue(row_number, "ERROR", None, key, "DUPLICATE_ROW", "文件内部存在重复记录"))
        seen_keys.add(key)
        severities = {problem["severity"] for problem in problems}
        if "ERROR" in severities:
            error_rows.add(row_number)
        else:
            valid += 1
        if "WARNING" in severities:
            warning_rows.add(row_number)
        for problem in problems:
            issue_counts[problem["issue_code"]] += 1
            db.add(ImportRowIssue(import_batch_id=batch.id, sheet_name=batch.selected_sheet_name or "", **problem))
    if batch.import_type == "WEEKLY_PLAN" and (batch.import_options or {}).get("plan_period_consistent") is False:
        warning_rows.add(0)
        problem = issue(0, "WARNING", "plan_start_date", None, "WEEKLY_PLAN_PERIOD_MISMATCH", "工作簿内不同工作表的计划周期不一致，请分别确认")
        issue_counts[problem["issue_code"]] += 1
        db.add(ImportRowIssue(import_batch_id=batch.id, sheet_name=batch.selected_sheet_name or "", **problem))
    batch.total_rows = total
    batch.valid_rows = valid
    batch.warning_rows = len(warning_rows)
    batch.error_rows = len(error_rows)
    batch.error_summary = dict(issue_counts)
    batch.import_options = {
        **(batch.import_options or {}),
        "hidden_data_rows": hidden_data_rows,
        "excluded_hidden_data_rows": excluded_hidden_data_rows,
    }
    batch.status = "VALIDATION_FAILED" if error_rows else "READY"
    db.flush()
    return {"total_rows": total, "valid_rows": valid, "warning_rows": len(warning_rows), "error_rows": len(error_rows), "hidden_data_rows": hidden_data_rows, "excluded_hidden_data_rows": excluded_hidden_data_rows, "status": batch.status}


def get_or_create_product(
    db: Session, normalized: dict[str, Any], batch: ImportBatch
) -> tuple[Product, dict[str, Any] | None]:
    product = db.scalar(select(Product).where(Product.product_code == normalized["product_code"]).with_for_update())
    incoming_fields = {
        "product_name": normalized.get("product_name"),
        "specification": normalized.get("specification"),
        "category": normalized.get("category"),
        "unit": normalized.get("unit"),
    }
    if product is None:
        product = Product(
            product_code=normalized["product_code"],
            **incoming_fields,
            data_source=batch.import_type,
            last_import_batch_id=batch.id,
        )
        db.add(product)
        db.flush()
        created_fields = ["product_code", *[field for field, value in incoming_fields.items() if value not in (None, "")]]
        return product, {
            "product_id": product.id,
            "change_type": "CREATED",
            "before": None,
            "after": {"product_code": product.product_code, **incoming_fields, "last_import_batch_id": batch.id},
            "fields": created_fields,
        }
    policy = str((batch.import_options or {}).get("master_data_policy", "FILL_EMPTY"))
    before = {**{field: getattr(product, field) for field in incoming_fields}, "last_import_batch_id": product.last_import_batch_id}
    changed_fields: list[str] = []
    for field, incoming_value in incoming_fields.items():
        if incoming_value in (None, ""):
            continue
        current_value = getattr(product, field)
        if policy == "FILL_EMPTY" and current_value in (None, ""):
            setattr(product, field, incoming_value)
            changed_fields.append(field)
        elif policy == "ADMIN_UPDATE" and current_value != incoming_value:
            setattr(product, field, incoming_value)
            changed_fields.append(field)
    if changed_fields:
        product.last_import_batch_id = batch.id
        db.flush()
        after = {**{field: getattr(product, field) for field in incoming_fields}, "last_import_batch_id": product.last_import_batch_id}
        return product, {
            "product_id": product.id,
            "change_type": "ADMIN_UPDATED" if policy == "ADMIN_UPDATE" else "FILLED_EMPTY",
            "before": before,
            "after": after,
            "fields": changed_fields,
        }
    return product, None


def build_record(batch: ImportBatch, row_number: int, normalized: dict[str, Any], product: Product):
    common = {
        "import_batch_id": batch.id,
        "source_sheet": batch.selected_sheet_name or "",
        "source_row_number": row_number,
        "raw_data": normalized["raw_data"],
        "product_id": product.id,
    }
    if batch.import_type == "SHIPMENT":
        return ShipmentRecord(**common, document_no=normalized["document_no"], shipment_date=normalized["shipment_date"], shipment_month=normalized.get("shipment_month"), quantity=normalized["quantity"], production_batch_no=normalized.get("production_batch_no"))
    if batch.import_type == "INVENTORY":
        return InventorySnapshot(**common, snapshot_date=normalized.get("snapshot_date"), on_hand_qty=normalized["on_hand_qty"], expected_inbound_qty=normalized["expected_inbound_qty"], expected_outbound_qty=normalized["expected_outbound_qty"], source_available_qty=normalized.get("source_available_qty"), calculated_available_qty=normalized["calculated_available_qty"])
    if batch.import_type == "PIPE_WIP":
        return PipeWipSnapshot(**common, snapshot_date=normalized.get("snapshot_date"), quantity=normalized["quantity"])
    if batch.import_type == "FITTING_WIP":
        return FittingWipSnapshot(**common, snapshot_date=normalized.get("snapshot_date"), production_batch_no=normalized.get("production_batch_no"), quantity=normalized["quantity"])
    if batch.import_type == "REGULAR_PRODUCT":
        return RegularProductionProduct(**common)
    return ImportedWeeklyPlanRaw(**common, production_batch_no=normalized["production_batch_no"], process_name=normalized["process_name"], equipment_name=normalized["equipment_name"], plan_start_date=normalized.get("plan_start_date"), plan_end_date=normalized.get("plan_end_date"), planned_quantity=normalized["planned_quantity"], actual_quantity=normalized.get("actual_quantity"), daily_plan=normalized.get("daily_plan", {}), daily_actual=normalized.get("daily_actual", {}))


RECORD_MODELS = (ShipmentRecord, InventorySnapshot, PipeWipSnapshot, FittingWipSnapshot, RegularProductionProduct, ImportedWeeklyPlanRaw, WeeklyPlanStagingRow)
PRODUCT_REFERENCE_MODELS = (ShipmentRecord, InventorySnapshot, PipeWipSnapshot, FittingWipSnapshot, RegularProductionProduct, ImportedWeeklyPlanRaw)


def import_validated_batch(db: Session, batch: ImportBatch, path: Path) -> tuple[int, list[dict[str, Any]]]:
    if batch.status != "IMPORTING" or batch.error_rows:
        raise ImportValidationError("IMPORT_VALIDATION_REQUIRED", "批次未通过全量校验或未由确认接口锁定")
    imported = 0
    product_changes_by_id: dict[int, dict[str, Any]] = {}
    parsed_rows = list(iter_normalized_rows(path, batch))
    if batch.import_type == "WEEKLY_PLAN":
        for row_number, normalized, problems in parsed_rows:
            if not normalized or any(item["severity"] == "ERROR" for item in problems):
                continue
            db.add(
                WeeklyPlanStagingRow(
                    import_batch_id=batch.id,
                    source_sheet=batch.selected_sheet_name or "",
                    source_row_number=row_number,
                    raw_data=normalized["raw_data"],
                    product_name_raw=normalized.get("product_name"),
                    specification_raw=normalized.get("specification"),
                    production_batch_no=normalized.get("production_batch_no"),
                    process_name=normalized.get("process_name"),
                    equipment_name=normalized.get("equipment_name"),
                    plan_start_date=normalized["plan_start_date"],
                    plan_end_date=normalized["plan_end_date"],
                    daily_plan=normalized["daily_plan"],
                    daily_actual=normalized["daily_actual"],
                    weekly_plan_qty=normalized["planned_quantity"],
                    weekly_actual_qty=normalized.get("actual_quantity"),
                    formula_metadata=normalized["formula_metadata"],
                    match_status="UNMATCHED",
                )
            )
            imported += 1
        db.flush()
        return imported, []
    for row_number, normalized, problems in parsed_rows:
        if not normalized or any(item["severity"] == "ERROR" for item in problems):
            continue
        product, product_change = get_or_create_product(db, normalized, batch)
        if product_change:
            existing_change = product_changes_by_id.get(product.id)
            if existing_change is None:
                product_changes_by_id[product.id] = product_change
            else:
                existing_change["after"] = product_change["after"]
                existing_change["fields"] = list(dict.fromkeys([*existing_change["fields"], *product_change["fields"]]))
                if existing_change["change_type"] != "CREATED" and product_change["change_type"] == "ADMIN_UPDATED":
                    existing_change["change_type"] = "ADMIN_UPDATED"
        db.add(build_record(batch, row_number, normalized, product))
        imported += 1
    for product_change in product_changes_by_id.values():
        db.add(
            ProductImportChange(
                import_batch_id=batch.id,
                product_id=product_change["product_id"],
                change_type=product_change["change_type"],
                before_data=product_change["before"],
                after_data=product_change["after"],
                changed_fields=product_change["fields"],
            )
        )
    db.flush()
    return imported, list(product_changes_by_id.values())


def delete_batch_records(db: Session, batch_id: int) -> None:
    for model in RECORD_MODELS:
        db.execute(delete(model).where(model.import_batch_id == batch_id))


def _product_has_other_batch_references(db: Session, product_id: int, batch_id: int) -> bool:
    for model in PRODUCT_REFERENCE_MODELS:
        reference = db.scalar(
            select(model.id).where(model.product_id == product_id, model.import_batch_id != batch_id).limit(1)
        )
        if reference is not None:
            return True
    staging_reference = db.scalar(
        select(WeeklyPlanStagingRow.id).where(
            WeeklyPlanStagingRow.matched_product_id == product_id,
            WeeklyPlanStagingRow.import_batch_id != batch_id,
        ).limit(1)
    )
    return staging_reference is not None


def rollback_product_import_changes(db: Session, batch: ImportBatch) -> list[dict[str, Any]]:
    changes = db.scalars(
        select(ProductImportChange)
        .where(ProductImportChange.import_batch_id == batch.id)
        .order_by(ProductImportChange.id.desc())
        .with_for_update()
    ).all()
    pending: list[tuple[ProductImportChange, Product]] = []
    for change in changes:
        product = db.scalar(select(Product).where(Product.id == change.product_id).with_for_update()) if change.product_id is not None else None
        if product is None:
            raise ImportValidationError(
                "PRODUCT_MASTER_ROLLBACK_CONFLICT",
                "产品主数据已不存在，不能安全撤销该导入批次",
                {"change_id": change.id},
            )
        if change.change_type == "CREATED":
            if product.last_import_batch_id != batch.id or _product_has_other_batch_references(db, product.id, batch.id):
                raise ImportValidationError(
                    "PRODUCT_MASTER_ROLLBACK_CONFLICT",
                    "当前批次新建的产品已被后续批次或其他数据引用，不能撤销",
                    {"product_id": product.id, "product_code": product.product_code},
                )
        elif product.last_import_batch_id != batch.id:
            raise ImportValidationError(
                "PRODUCT_MASTER_ROLLBACK_CONFLICT",
                "产品字段在当前批次之后又被其他批次修改，不能自动恢复",
                {"product_id": product.id, "product_code": product.product_code, "last_import_batch_id": product.last_import_batch_id},
            )
        pending.append((change, product))

    delete_batch_records(db, batch.id)
    restoration_events: list[dict[str, Any]] = []
    for change, product in pending:
        before_restore = {
            "product_code": product.product_code,
            "product_name": product.product_name,
            "specification": product.specification,
            "category": product.category,
            "unit": product.unit,
            "last_import_batch_id": product.last_import_batch_id,
        }
        if change.change_type == "CREATED":
            db.delete(product)
            after_restore = None
        else:
            original = change.before_data or {}
            for field in change.changed_fields:
                if field in {"product_name", "specification", "category", "unit"}:
                    setattr(product, field, original.get(field))
            product.last_import_batch_id = original.get("last_import_batch_id")
            after_restore = {
                "product_code": product.product_code,
                "product_name": product.product_name,
                "specification": product.specification,
                "category": product.category,
                "unit": product.unit,
                "last_import_batch_id": product.last_import_batch_id,
            }
        restoration_events.append(
            {
                "product_id": change.product_id,
                "product_code": before_restore["product_code"],
                "change_type": change.change_type,
                "before": before_restore,
                "after": after_restore,
                "changed_fields": change.changed_fields,
            }
        )
    db.flush()
    return restoration_events


def batch_has_downstream_references(db: Session, batch: ImportBatch) -> bool:
    # Stage 2 has no downstream production models. Future stages extend this guard
    # with explicit foreign-key reference checks before allowing rollback.
    return False
